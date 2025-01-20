import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# 엑셀 데이터베이스 초기화
file_name = "수입_지출_관리.xlsx"
try:
    wb = openpyxl.load_workbook(file_name)
    print("기존 엑셀 파일을 불러왔습니다.")
except FileNotFoundError:
    wb = openpyxl.Workbook()

    # 기본 시트 생성
    income_sheet = wb.active
    income_sheet.title = "수입관리"
    income_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
    income_sheet.append(income_headers)

    expense_sheet = wb.create_sheet(title="지출관리")
    expense_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
    expense_sheet.append(expense_headers)

    summary_sheet = wb.create_sheet(title="월별요약")
    summary_headers = ["월", "총수입 (면세)", "총수입 (과세)", "총지출 (면세)", "총지출 (과세)", "순이익"]
    summary_sheet.append(summary_headers)

    annual_sheet = wb.create_sheet(title="연간요약")
    annual_headers = ["년도", "총수입 (면세)", "총수입 (과세)", "총지출 (면세)", "총지출 (과세)", "순이익"]
    annual_sheet.append(annual_headers)

    wb.save(file_name)
    print("새로운 엑셀 파일을 생성했습니다.")

# 월별 및 연간 데이터 갱신 함수
def update_summary():
    income_sheet = wb["수입관리"]
    expense_sheet = wb["지출관리"]
    summary_sheet = wb["월별요약"]
    annual_sheet = wb["연간요약"]

    # 초기화
    summary_sheet.delete_rows(2, summary_sheet.max_row)
    annual_sheet.delete_rows(2, annual_sheet.max_row)

    monthly_data = {}
    annual_data = {}

    # 데이터 집계 함수
    def process_row(sheet, row, is_income):
        date = datetime.strptime(row[0], "%Y-%m-%d")
        year = date.year
        month = date.strftime("%Y-%m")
        category = row[2]
        amount = float(row[3])

        if month not in monthly_data:
            monthly_data[month] = {"수입_면세": 0, "수입_과세": 0, "지출_면세": 0, "지출_과세": 0}

        if year not in annual_data:
            annual_data[year] = {"수입_면세": 0, "수입_과세": 0, "지출_면세": 0, "지출_과세": 0}

        key = ("수입_" if is_income else "지출_") + category
        monthly_data[month][key] += amount
        annual_data[year][key] += amount

    # 수입 및 지출 데이터 처리
    for row in income_sheet.iter_rows(min_row=2, values_only=True):
        process_row(income_sheet, row, is_income=True)

    for row in expense_sheet.iter_rows(min_row=2, values_only=True):
        process_row(expense_sheet, row, is_income=False)

    # 월별 데이터 기록
    for month, data in sorted(monthly_data.items()):
        net_income = (data["수입_면세"] + data["수입_과세"]) - (data["지출_면세"] + data["지출_과세"])
        summary_sheet.append([month, data["수입_면세"], data["수입_과세"], data["지출_면세"], data["지출_과세"], net_income])

    # 연간 데이터 기록
    for year, data in sorted(annual_data.items()):
        net_income = (data["수입_면세"] + data["수입_과세"]) - (data["지출_면세"] + data["지출_과세"])
        annual_sheet.append([year, data["수입_면세"], data["수입_과세"], data["지출_면세"], data["지출_과세"], net_income])

    wb.save(file_name)

# 데이터 추가 함수
def add_data():
    def save_data():
        date = entry_date.get()
        client = entry_client.get()
        category = combo_category.get()
        amount = entry_amount.get()
        note = entry_note.get()
        sheet_name = combo_sheet.get()

        if date and client and category and amount and sheet_name:
            sheet = wb[sheet_name]
            sheet.append([date, client, category, float(amount), note])
            wb.save(file_name)
            update_summary()
            messagebox.showinfo("데이터 입력", "데이터가 성공적으로 추가되었습니다!")
            add_window.destroy()
        else:
            messagebox.showwarning("입력 오류", "모든 필드를 정확히 입력해주세요.")

    add_window = tk.Toplevel()
    add_window.title("데이터 추가")
    add_window.geometry("400x400")

    tk.Label(add_window, text="날짜 (YYYY-MM-DD):").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    entry_date = tk.Entry(add_window)
    entry_date.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(add_window, text="거래처:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    entry_client = tk.Entry(add_window)
    entry_client.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(add_window, text="항목 (면세/과세):").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    combo_category = ttk.Combobox(add_window, values=["면세", "과세"], state="readonly")
    combo_category.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(add_window, text="금액:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
    entry_amount = tk.Entry(add_window)
    entry_amount.grid(row=3, column=1, padx=10, pady=5)

    tk.Label(add_window, text="비고:").grid(row=4, column=0, padx=10, pady=5, sticky="e")
    entry_note = tk.Entry(add_window)
    entry_note.grid(row=4, column=1, padx=10, pady=5)

    tk.Label(add_window, text="수입/지출:").grid(row=5, column=0, padx=10, pady=5, sticky="e")
    combo_sheet = ttk.Combobox(add_window, values=["수입관리", "지출관리"], state="readonly")
    combo_sheet.grid(row=5, column=1, padx=10, pady=5)

    btn_save = tk.Button(add_window, text="저장", command=save_data)
    btn_save.grid(row=6, column=0, columnspan=2, pady=10)

# 데이터 확인 함수
def view_data(sheet_name):
    view_window = tk.Toplevel()
    view_window.title(f"{sheet_name} 데이터 확인")
    view_window.geometry("800x400")

    columns = [cell.value for cell in wb[sheet_name][1]]
    tree = ttk.Treeview(view_window, columns=columns, show="headings")
    tree.pack(fill="both", expand=True)

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100, anchor="center")

    for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)


# GUI 생성
root = tk.Tk()
root.title("수입/지출 관리 프로그램")
root.geometry("600x500")

# 버튼 추가
header_label = tk.Label(root, text="수입/지출 관리 프로그램", font=("Arial", 16))
header_label.pack(pady=10)

# 데이터 추가 버튼
data_add_button = tk.Button(root, text="수입/지출 추가", width=20, command=add_data)
data_add_button.pack(pady=10)

# 수입 데이터 확인 버튼
view_income_button = tk.Button(root, text="수입 데이터 확인", width=20, command=lambda: view_data("수입관리"))
view_income_button.pack(pady=10)

# 지출 데이터 확인 버튼
view_expense_button = tk.Button(root, text="지출 데이터 확인", width=20, command=lambda: view_data("지출관리"))
view_expense_button.pack(pady=10)

# 월별 요약 확인 버튼
view_monthly_summary_button = tk.Button(root, text="월별 요약 확인", width=20, command=lambda: view_data("월별요약"))
view_monthly_summary_button.pack(pady=10)

# 연간 요약 확인 버튼
view_annual_summary_button = tk.Button(root, text="연간 요약 확인", width=20, command=lambda: view_data("연간요약"))
view_annual_summary_button.pack(pady=10)

# 종료 버튼
exit_button = tk.Button(root, text="종료", width=20, command=root.destroy)
exit_button.pack(pady=20)

# 프로그램 시작 시 요약 데이터 갱신
update_summary()

# 메인 루프 시작
root.mainloop()
