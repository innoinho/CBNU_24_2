import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# 새 엑셀 워크북 생성
wb = openpyxl.Workbook()

# 1. 수입 관리 시트 생성
income_sheet = wb.active
income_sheet.title = "수입관리"

# 수입 관리 시트 헤더 추가
income_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
for col_num, header in enumerate(income_headers, 1):
    cell = income_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 샘플 데이터 추가 (수입 관리)
income_sample_data = [
    ["2024-01-01", "거래처 A", "면세", 100000, "비고 1"],
    ["2024-01-02", "거래처 B", "과세", 200000, "비고 2"],
    ["2024-01-03", "거래처 C", "면세", 150000, "비고 3"],
]
for row_num, row_data in enumerate(income_sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        income_sheet.cell(row=row_num, column=col_num, value=value)

# 2. 지출 관리 시트 생성
expense_sheet = wb.create_sheet(title="지출관리")

# 지출 관리 시트 헤더 추가
expense_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
for col_num, header in enumerate(expense_headers, 1):
    cell = expense_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 샘플 데이터 추가 (지출 관리)
expense_sample_data = [
    ["2024-01-01", "거래처 D", "면세", 50000, "비고 1"],
    ["2024-01-02", "거래처 E", "과세", 80000, "비고 2"],
    ["2024-01-03", "거래처 F", "면세", 60000, "비고 3"],
]
for row_num, row_data in enumerate(expense_sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        expense_sheet.cell(row=row_num, column=col_num, value=value)

# 3. 월별 요약 시트 생성
monthly_summary_sheet = wb.create_sheet(title="월별요약")

# 월별 요약 시트 헤더 추가
monthly_headers = ["월", "총 수입 (면세)", "총 수입 (과세)", "총 지출 (면세)", "총 지출 (과세)", "순이익"]
for col_num, header in enumerate(monthly_headers, 1):
    cell = monthly_summary_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 월별 요약 계산식 추가
monthly_summary_sheet.cell(row=2, column=1, value="1월")
monthly_summary_sheet.cell(row=2, column=2, value="=SUMIF(수입관리!C:C, \"면세\", 수입관리!D:D)")
monthly_summary_sheet.cell(row=2, column=3, value="=SUMIF(수입관리!C:C, \"과세\", 수입관리!D:D)")
monthly_summary_sheet.cell(row=2, column=4, value="=SUMIF(지출관리!C:C, \"면세\", 지출관리!D:D)")
monthly_summary_sheet.cell(row=2, column=5, value="=SUMIF(지출관리!C:C, \"과세\", 지출관리!D:D)")
monthly_summary_sheet.cell(row=2, column=6, value="=B2+C2-D2-E2")

# 4. 연간 요약 시트 생성
annual_summary_sheet = wb.create_sheet(title="연간요약")

# 연간 요약 시트 헤더 추가
annual_headers = ["연도", "총 수입 (면세)", "총 수입 (과세)", "총 지출 (면세)", "총 지출 (과세)", "순이익"]
for col_num, header in enumerate(annual_headers, 1):
    cell = annual_summary_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 연간 요약 계산식 추가
annual_summary_sheet.cell(row=2, column=1, value="2024")
annual_summary_sheet.cell(row=2, column=2, value="=SUM(월별요약!B2:B13)")
annual_summary_sheet.cell(row=2, column=3, value="=SUM(월별요약!C2:C13)")
annual_summary_sheet.cell(row=2, column=4, value="=SUM(월별요약!D2:D13)")
annual_summary_sheet.cell(row=2, column=5, value="=SUM(월별요약!E2:E13)")
annual_summary_sheet.cell(row=2, column=6, value="=B2+C2-D2-E2")

# 열 너비 조정
for sheet in [income_sheet, expense_sheet, monthly_summary_sheet, annual_summary_sheet]:
    for col_num in range(1, len(income_headers) + 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

# 파일 저장
file_name = "회계_수입_지출_관리_양식.xlsx"
wb.save(file_name)
print(f"엑셀 파일 '{file_name}'이 생성되었습니다.")
