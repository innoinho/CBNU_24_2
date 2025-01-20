import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# 새 엑셀 워크북 생성
wb = openpyxl.Workbook()

# 1. 수입 관리 시트 생성
income_sheet = wb.active
income_sheet.title = "수입관리"

# 수입 관리 시트 헤더 추가
income_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
for col_num, header in enumerate(income_headers, 1):
    cell = income_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 2. 지출 관리 시트 생성
expense_sheet = wb.create_sheet(title="지출관리")

# 지출 관리 시트 헤더 추가
expense_headers = ["날짜", "거래처", "항목 (면세/과세 구분)", "금액", "비고"]
for col_num, header in enumerate(expense_headers, 1):
    cell = expense_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 3. 월별 요약 시트 생성
monthly_summary_sheet = wb.create_sheet(title="월별요약")

# 월별 요약 시트 헤더 추가
monthly_headers = ["월", "총 수입 (면세)", "총 수입 (과세)", "총 지출 (면세)", "총 지출 (과세)", "순이익"]
for col_num, header in enumerate(monthly_headers, 1):
    cell = monthly_summary_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 4. 연간 요약 시트 생성
annual_summary_sheet = wb.create_sheet(title="연간요약")

# 연간 요약 시트 헤더 추가
annual_headers = ["연도", "총 수입 (면세)", "총 수입 (과세)", "총 지출 (면세)", "총 지출 (과세)", "순이익"]
for col_num, header in enumerate(annual_headers, 1):
    cell = annual_summary_sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 열 너비 조정
for sheet in [income_sheet, expense_sheet, monthly_summary_sheet, annual_summary_sheet]:
    for col_num in range(1, len(income_headers) + 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

# 파일 저장
file_name = "수입_지출_관리.xlsx"
wb.save(file_name)
print(f"엑셀 파일 '{file_name}'이 생성되었습니다.")